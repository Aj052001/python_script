import sys, os, requests
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from prometheus_api_client import PrometheusConnect
from collections import defaultdict
import datetime as dt

PROMETHEUS_URL = os.getenv("PROMETHEUS_URL", "http://localhost:9090")
NAMESPACE = "default"
STEP = "60s"

def load_deployments_from_excel(filepath: str):
    deployments = []
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    # Try header "Deployment" else first column
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return deployments
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    try:
        col_idx = header.index("deployment")
        start = 1
    except ValueError:
        col_idx = 0
        start = 1  # skip header anyway
    for r in rows[start:]:
        if r and r[col_idx]:
            deployments.append(str(r[col_idx]).strip())
    return deployments

def get_pods_of_deployment(deployment_name):
    query = f'series?match[]=container_cpu_usage_seconds_total{{namespace="{NAMESPACE}",pod=~"{deployment_name}-.*"}}'
    try:
        response = requests.get(f"{PROMETHEUS_URL}/api/v1/{query}", timeout=30)
        response.raise_for_status()
        data = response.json()
        pods = set()
        for item in data.get("data", []):
            pod_name = item.get("pod")
            if pod_name:
                pods.add(pod_name)
        return list(pods)
    except Exception as e:
        print("Error fetching pod list:", e)
        return []

def run_prometheus_query(query):
    end_time = datetime.now()
    start_time = end_time - timedelta(days=7)
    step = 60
    try:
        response = requests.get(
            f"{PROMETHEUS_URL}/api/v1/query_range",
            params={
                "query": query, "start": start_time.timestamp(),
                "end": end_time.timestamp(), "step": step,
            },
            timeout=60,
        )
        if not response.ok:
            print("QUERY:\n", query)
            print("STATUS:", response.status_code)
            print("BODY:", response.text)
            response.raise_for_status()
        result = response.json().get("data", {}).get("result", [])
        values_all = []
        for item in result:
            values_all.extend(float(v[1]) for v in item.get("values", []))
        if not values_all:
            return 0, 0
        return max(values_all), (sum(values_all) / len(values_all))
    except Exception as e:
        print("Error running query:", e)
        return 0, 0

def get_cpu_stats_for_pod(pod_name: str):
    query = f"""
sum(rate(container_cpu_usage_seconds_total{{namespace="{NAMESPACE}", pod="{pod_name}"}}[5m])) by (pod,instance) * 1000 * 100 * 100
/ sum(container_spec_cpu_quota{{namespace="{NAMESPACE}", pod="{pod_name}"}}) by (pod,instance)
"""
    return run_prometheus_query(query)

def get_memory_stats_for_deployment(deployment_name):
    try:
        prom = PrometheusConnect(url=PROMETHEUS_URL, disable_ssl=True)
        end_time = dt.datetime.now(dt.timezone.utc).replace(microsecond=0)
        start_time = end_time - dt.timedelta(days=7)
        query = f"""
            (sum(container_memory_working_set_bytes{{pod=~"{deployment_name}-.*", namespace="{NAMESPACE}", container!="POD"}}) by (pod)
            /
            sum(container_spec_memory_limit_bytes{{pod=~"{deployment_name}-.*", namespace="{NAMESPACE}", container!="POD"}}) by (pod)) * 100
        """
        memory_data = prom.custom_query_range(
            query=query, start_time=start_time, end_time=end_time, step=STEP
        )
        if not memory_data:
            return 0, 0
        all_values = [float(v[1]) for item in memory_data for v in item['values']]
        max_usage = max(all_values) if all_values else 0
        daily_values = defaultdict(list)
        for item in memory_data:
            for ts_str, val_str in item['values']:
                dt_obj = dt.datetime.fromtimestamp(float(ts_str), dt.timezone.utc)
                daily_values[dt_obj.date()].append(float(val_str))
        daily_means = [sum(v)/len(v) for v in daily_values.values() if v]
        max_of_mean_daily_usage = max(daily_means) if daily_means else 0
        return round(max_usage, 2), round(max_of_mean_daily_usage, 2)
    except Exception as e:
        print(f"Memory query error for {deployment_name}: {e}")
        return 0, 0

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python pod_stats.py <deployments_excel.xlsx>")
        sys.exit(1)

    deployments_file = sys.argv[1]
    DEPLOYMENTS = load_deployments_from_excel(deployments_file)
    if not DEPLOYMENTS:
        print("❌ No deployments found in Excel. Exiting...")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pod Stats"
    ws.append(["Deployment", "Max_CPU (%)", "Mean_CPU (%)", "Max_Memory (%)", "Max_of_Daily_Mean_Memory (%)"])

    for DEPLOYMENT_NAME in DEPLOYMENTS:
        pods = get_pods_of_deployment(DEPLOYMENT_NAME)
        if not pods:
            print(f"⚠️ No pods found for {DEPLOYMENT_NAME}")
            continue

        max_cpu_overall, max_mean_cpu_overall = 0, 0
        for pod in pods:
            pod_max_cpu, pod_mean_cpu = get_cpu_stats_for_pod(pod)
            if pod_max_cpu > max_cpu_overall:
                max_cpu_overall = pod_max_cpu
            if pod_mean_cpu > max_mean_cpu_overall:
                max_mean_cpu_overall = pod_mean_cpu

        max_mem, max_mean_mem = get_memory_stats_for_deployment(DEPLOYMENT_NAME)

        print(f"\n⚡ {DEPLOYMENT_NAME}:")
        print(f"CPU → Max: {max_cpu_overall:.2f}% | Mean: {max_mean_cpu_overall:.2f}%")
        print(f"Memory → Max: {max_mem:.2f}% | Max of Daily Mean: {max_mean_mem:.2f}%")

        ws.append([
            DEPLOYMENT_NAME,
            round(max_cpu_overall, 2),
            round(max_mean_cpu_overall, 2),
            max_mem,
            max_mean_mem
        ])

    filename = "pod_summary_with_memory.xlsx"
    wb.save(filename)
    print(f"\n📊 Summary saved to {filename}")
